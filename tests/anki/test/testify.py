from pyquery import PyQuery
import re
import json
import os

from supermemo_toolkit.utilscripts.ankinet import invoke
import ast
from openai import OpenAI
import urllib.request
from bs4 import BeautifulSoup
from pyquery import PyQuery as pq

import openpyxl
import spacy  # 导包
from nltk.stem import PorterStemmer
# import html
# import json
# import urllib.request
# from pyquery import PyQuery

a = [
    '<span class="definition">vt. ',
    "相关词义延伸：\n",
    "· pool可指从不同的人那里收集。\n",
]
# b = list()
# for item in a:
#     t = re.search(r"class", item)
#     if t is not None:
#         b.append(item)
# print(b)

# 不能这样做
# for item in a:
#     t = re.search(r"class", item)
#     if t is None:
#         a.remove(item)
#         print(t)
# print(a)
explain_words = str()
for item in a:
    explain_words = explain_words + "<b></b>" + str(item)
print(explain_words)


# ————————————————


html = "<div class='definitions'> <div class='hom'><span class='gramGrp pos'>uncountable noun</span><div class='sense'> <div class='def'><span class='hi rend-b'>Anguish</span> is <a class='ref type-def' href='entry://great'>great</a> mental <a class='ref type-def' href='entry://suffering'>suffering</a> or physical pain. <span class='tips'><span class='lbl type-register'><span class='punctuation'> [</span>written<span class='punctuation'>]</span></span></span></div><div class='examples'><div class='cit'><span class='quote'> A cry of anguish burst from her lips.</span><span class='type-exa_sound'> <span class='exa_sound online' data-lang='en_GB' data-src-mp3='https://www.collinsdictionary.com/sounds/hwd_sounds/en_gb_exa_a_cry_of_anguish_burst_from_her_lips.mp3' onclick='new Audio(this.getAttribute('data-src-mp3')).play()'> </span></span></div><div class='cit'><span class='quote'> Mark looked at him in anguish.</span><span class='type-exa_sound'> <span class='exa_sound online' data-lang='en_GB' data-src-mp3='https://www.collinsdictionary.com/sounds/hwd_sounds/en_gb_exa_mark_looked_at_him_in_anguish.mp3' onclick='new Audio(this.getAttribute('data-src-mp3')).play()'> </span></span></div></div> <div class='thes'><b>Synonyms: </b> <a class='form ref' href='entry://suffering'>suffering</a>, <a class='form ref' href='entry://pain'>pain</a>, <a class='form ref' href='entry://torture'>torture</a>, <a class='form ref' href='entry://distress'>distress</a> &nbsp; </div> </div> </div> </div>"

pqdoc = PyQuery(html)
# explain_entrys = pqdoc('div > div.cobuild > div.definitions')
for a_tag in pqdoc("div.def > a"):
    # 创建一个新的<u>标签，并将<a>标签的文本内容赋值给它
    span_tag = PyQuery("<span></span>")
    span_tag.attr["class"] = a_tag.attrib["class"]
    span_tag.text(a_tag.text)
    pqdoc(a_tag).replaceWith(span_tag)
explain = str(pqdoc)
print(explain)


# # with open("C:/Users/Snowy/Desktop/adveng 2018 (all elements).xml", "r") as f:
# #     tree = xmltodict.parse(f.read())

# a = html.unescape("r&amp;#618;&amp;#39;me&amp;#618;nd&amp;#601;&amp;#691;")
# print(a)
# tree = etree.parse("C:/Users/Snowy/Desktop/adveng 2018 (all elements).xml")
# root = tree.getroot()
# # 查找所有的<Content>元素
# content_elements = root.xpath("//Content")

# # 筛选出同时包含<Question>和<Answer>子元素的<Content>元素
# filtered_elements = [content for content in content_elements if content.xpath(".//Question") and content.xpath(".//Answer")]
# for element in filtered_elements:
#     print(etree.tostring(element, pretty_print=True).decode())
# pass

# # 必须有Question和Answer的Content才行

# ——————————————————————


# note_id_list = invoke(
#     "findNotes", query="deck:2024红宝书考研词汇"
# )
# notesInfo = invoke("notesInfo", notes=note_id_list)
# mjson = dict()
# for index, noteInfo in enumerate(notesInfo):
#     noteId = noteInfo["noteId"]
#     tags: list = noteInfo["tags"]
#     fields = noteInfo["fields"]
#     word = fields["word"]["value"]
#     tyj = fields["同义句"]["value"]
#     tyj_mean = fields["同义句释义"]["value"]
#     mjson[word] = [tyj, tyj_mean]
#     print(word, f"{index + 1}/{len(notesInfo)}")
#
# with open(os.path.join(os.getcwd(),"m.json"), "w", encoding="utf-8") as json_file:
#     json.dump(mjson, json_file, indent=4)


with open(os.path.join(os.getcwd(), "m.json"), "rb") as f:
    raw_data = f.read()
    json_data = json.loads(raw_data)

note_id_list = invoke("findNotes", query="deck:24意境语义红宝石")
notesInfo = invoke("notesInfo", notes=note_id_list)
for index, noteInfo in enumerate(notesInfo):
    noteId = noteInfo["noteId"]
    tags: list = noteInfo["tags"]
    fields = noteInfo["fields"]
    word = fields["word"]["value"]
    if word in json_data:
        invoke(
            "updateNote",
            note={
                "id": noteId,
                "fields": {
                    "同义短句": json_data[word][0],
                    "同义释义": json_data[word][1],
                },
            },
        )
        print(fields["word"]["value"], f"{index + 1}/{len(notesInfo)}")

# ——————————————————————


note_id_list = invoke("findNotes", query="deck:2024红宝书考研词汇")
notesInfo = invoke("notesInfo", notes=note_id_list)
for index, noteInfo in enumerate(notesInfo):
    noteId = noteInfo["noteId"]
    tags: list = noteInfo["tags"]
    fields = noteInfo["fields"]
    try:
        TJF = json.loads(fields["同近反"]["value"].replace("'", '"'))
        if "related" in TJF:
            del TJF["related"]
            result = json.dumps(TJF, ensure_ascii=False)
            invoke(
                "updateNote",
                note={"id": noteId, "fields": {"同近反": result}},
            )
            print(fields["word"]["value"], f"{index + 1}/{len(notesInfo)}")
        else:
            continue
    except Exception as e:
        print(e)


# ——————————————————————


def request(action, **params):
    return {"action": action, "params": params, "version": 6}


def invoke(action, **params):
    requestJson = json.dumps(request(action, **params)).encode("utf-8")
    response = json.load(
        urllib.request.urlopen(
            urllib.request.Request("http://127.0.0.1:8765", requestJson)
        )
    )
    if len(response) != 2:
        raise Exception("response has an unexpected number of fields")
    if "error" not in response:
        raise Exception("response is missing required error field")
    if "result" not in response:
        raise Exception("response is missing required result field")
    if response["error"] is not None:
        raise Exception(response["error"])
    return response["result"]


client = OpenAI(
    # #将这里换成你在aihubmix api keys拿到的密钥
    api_key="sk-EonFLRSFSs0w8Llt4cA95aB8529d47D884Ac0aC684C4084f",
    # 这里将官方的接口访问地址，替换成aihubmix的入口地址
    base_url="https://aihubmix.com/v1",
)


# Create a simple, easy-to-understand English phrase that encapsulates the essence of the given word without using the word itself, making sure a 6-year-old child in the U.S. can comprehend it. Provide only two parts: the phrase and its Chinese meaning, separated by a line break.\n\nExample:\nPhrase: Make code run\nMeaning: 让代码运行起来
def get_words(word: str):
    response = client.chat.completions.create(
        model="gpt-3.5-turbo-0125",  # 填写需要调用的模型名称gpt-3.5-turbo-0125
        messages=[
            {
                "role": "user",
                # "content": f"Find the closest synonyms, related words, and antonyms for the word '{word}'. Sort the results by relevance. Provide the output in a structured format, for example: {{'synonyms': ['synonym1', 'synonym2'], 'related': ['related1', 'related2'], 'antonyms': ['antonym1', 'antonym2']}}",
                "content": "You are an expert in English vocabulary for postgraduate entrance exams in China."
                + "Please prioritize words that are frequently tested and expected to be known by candidates appearing for postgraduate entrance exams in China."
                + f"Please provide the closest synonyms, related words, and antonyms for the word '{word}'."
                + "Sort the results by relevance. Provide the output in a structured format, for example: {'synonyms': ['synonym1', 'synonym2'], 'related': ['related1', 'related2'], 'antonyms': ['antonym1', 'antonym2']}",
            }
        ],
    )

    try:
        response_content = response.choices[0].message.content
        # 将响应内容转换为字典
        word_relationships = ast.literal_eval(response_content)
        return word_relationships
    except Exception as e:
        print(f"Error parsing response: {e}")
        return None


note_id_list = invoke("findNotes", query="deck:2024红宝书考研词汇")
notesInfo = invoke("notesInfo", notes=note_id_list)

for index, noteInfo in enumerate(notesInfo):
    noteId = noteInfo["noteId"]
    tags: list = noteInfo["tags"]
    fields = noteInfo["fields"]
    word = fields["word"]["value"]
    anki_tjf = fields["同近反"]["value"]
    if anki_tjf == "":
        tjf = get_words(word)
        if tjf:
            filed_tjf = str(tjf)
            tags.append("ky-tjf")
            invoke(
                "updateNote",
                note={
                    "id": noteId,
                    "fields": {"同近反": filed_tjf},
                    "tags": tags,
                },
            )
            print("同近反", f"{index + 1}/{len(notesInfo)}")

# ——————————————————————

# note_id_list = invoke(
#     "findNotes", query="deck:2024红宝书考研词汇（必考词+基础词+超纲词）"
# )
# notesInfo = invoke("notesInfo", notes=note_id_list)
# for index, noteInfo in enumerate(notesInfo):
#     noteId = noteInfo["noteId"]
#     tags: list = noteInfo["tags"]
#     fields = noteInfo["fields"]
#     if "剑桥LE" not in tags:
#         explain = download_cambridge_words_explain(fields["word"]["value"])
#         if explain:
#             invoke(
#                 "updateNote",
#                 note={
#                     "id": noteId,
#                     "fields": {"剑桥LE英英释义": explain},
#                     "tags": ["剑桥LE"],
#                 },
#             )
#             print(fields["word"]["value"], "explain", f"{index+1}/{len(notesInfo)}")

# <br>
# with open("./mdict.json", "rb") as f:
#     raw_data = f.read()
#     mdict = json.loads(raw_data)

# note_id_list = invoke(
#     "findNotes", query="deck:2024红宝书考研词汇（必考词+基础词+超纲词）"
# )
# notesInfo = invoke("notesInfo", notes=note_id_list)
# for noteInfo in notesInfo:
#     noteId = noteInfo["noteId"]
#     tags: list = noteInfo["tags"]
#     fields = noteInfo["fields"]
#     word: str = fields["word"]["value"]
#     if word in mdict.keys():
#         if len(mdict[word]) >= 1:
#             more_meaning = mdict[word]
#             cpath = list()
#             for single in more_meaning:
#                 if len(single) >= 1:
#                     cpath.append(single[0])
#             cyfl_path = "<br>".join(cpath)
#             invoke(
#                 "updateNote",
#                 note={
#                     "id": noteId,
#                     "fields": {"词义分类": cyfl_path},
#                     "tags": [""],
#                 },
#             )
#             print(noteId, cyfl_path, word)
# pass

# note_id_list1 = invoke(
#     "findNotes", query="deck:2024红宝书考研词汇（必考词+基础词+超纲词）"
# )
# notesInfo1 = invoke("notesInfo", notes=note_id_list1)

# words = dict()
# for noteInfo in notesInfo1:
#     words.update({noteInfo["fields"]["word"]["value"]: noteInfo["noteId"]})

# >>>
# note_id_list2 = invoke("findNotes", query="deck:Anki创享岛-2024恋练有词")
# notesInfo2 = invoke("notesInfo", notes=note_id_list2)
# wait_list = dict()
# for index, noteInfo in enumerate(notesInfo2):
#     noteId = noteInfo["noteId"]
#     tags: list = noteInfo["tags"]
#     fields = noteInfo["fields"]
#     if fields["英"]["value"] not in words:
#         yin = fields["音"]["value"]
#         yi = fields["译"]["value"]
#         zu = fields["组"]["value"].replace(" ","-")
#         wait_list.update({fields["英"]["value"]: (yin, yi, zu)})

# for word, tup in wait_list.items():
#     (yin, yi, zu) = tup
#     invoke(
#         "addNote",
#         note={
#             "deckName": "2024红宝书考研词汇（必考词+基础词+超纲词）::D - 扩展词",
#             "modelName": "新模板 6.0",
#             "fields": {
#                 "word": word,
#                 "单词": word,
#                 "US": yin,
#                 "释义": yi,
#             },
#             "options": {
#                 "allowDuplicate": False,
#                 "duplicateScope": "deck",
#                 "duplicateScopeOptions": {
#                     "deckName": "Default",
#                     "checkChildren": False,
#                     "checkAllModels": False,
#                 },
#             },
#             "tags": ["2024恋练有词", zu],
#         },
#     )
#     print(word, "add")


# pass
# >>>
# note_id_list = invoke(
#     "findNotes", query="deck:2024红宝书考研词汇（必考词+基础词+超纲词）"
# )
# notesInfo = invoke("notesInfo", notes=note_id_list)
# for index, noteInfo in enumerate(notesInfo):
#     noteId = noteInfo["noteId"]
#     tags: list = noteInfo["tags"]
#     fields = noteInfo["fields"]
#     if "剑桥LE" not in tags:
#         explain = download_cambridge_words_explain(fields["word"]["value"])
#         if explain:
#             invoke(
#                 "updateNote",
#                 note={
#                     "id": noteId,
#                     "fields": {"剑桥LE英英释义": explain},
#                     "tags": ["剑桥LE"],
#                 },
#             )
#             print(fields["word"]["value"], "explain", f"{index+1}/{len(notesInfo)}")


# >>> download_collins_words_explain
# note_id_list = invoke("findNotes", query="deck:2024红宝书考研词汇")
# notesInfo = invoke("notesInfo", notes=note_id_list)

# mdx = MDX("D:\\Software\\MDictPC\\doc\\Collins\ccald9.mdx")
# headwords = [*mdx]  # 单词名列表
# items = [*mdx.items()]  # 释义html源码列表
# if len(headwords) == len(items):
#     print(f"加载成功：共{len(headwords)}条")
# else:
#     print(f"【ERROR】加载失败{len(headwords)}，{len(items)}")

# for index, noteInfo in enumerate(notesInfo):
#     noteId = noteInfo["noteId"]
#     tags: list = noteInfo["tags"]
#     fields = noteInfo["fields"]
#     words = fields["word"]["value"]
#     if words.encode() in headwords:
#         word, html = items[headwords.index(words.encode())]
#         word, html = word.decode(), html.decode()
#         pqdoc = PyQuery(html)
#         explain_entrys = pqdoc("div > div.cobuild > div.definitions")
#         for a_tag in explain_entrys("div.def > a"):
#             # 创建一个新的<u>标签，并将<a>标签的文本内容赋值给它
#             span_tag = PyQuery("<span></span>")
#             span_tag.attr["class"] = a_tag.attrib["class"]
#             span_tag.text(a_tag.text)
#             explain_entrys(a_tag).replaceWith(span_tag)
#         a_hwd_sound_mdd = explain_entrys("a.hwd_sound")
#         a_tag_html = [
#             PyQuery(a_tag).html()
#             for a_tag in a_hwd_sound_mdd
#             if PyQuery(a_tag).html() != ""
#         ]
#         for a_tag in a_hwd_sound_mdd:
#             if PyQuery(a_tag).html() == "":
#                 a_hwd_sound_mdd(a_tag).remove()
#         # 如果a标签没有内容就啥也不做，反之去掉标签a。
#         if len(a_tag_html) > 0:
#             for index, a_tag in enumerate(a_hwd_sound_mdd):
#                 sub = a_tag_html[index]
#                 explain_entrys(a_tag).replaceWith(PyQuery(sub))
#             # PyQuery有定义__str__魔法方法，当试图将一个对象转换为字符串时会调用此方法。
#         explain = str(explain_entrys)
#         invoke(
#             "updateNote",
#             note={
#                 "id": noteId,
#                 "fields": {"柯林斯": explain},
#                 # "tags": tags,
#             },
#         )
#         print(fields["word"]["value"], "explain", f"{index+1}/{len(notesInfo)}")
#     else:
#         print(f"【未找到单词】：{words}")
#     # if "柯林斯" not in tags:
#     # tags.append("柯林斯")
#     # if "Mdict柯林斯" in tags:
#     #     tags.remove("Mdict柯林斯")

# >>>
# note_id_list = invoke("findNotes", query="deck:2024红宝书考研词汇")
# notesInfo = invoke("notesInfo", notes=note_id_list)

# for index, noteInfo in enumerate(notesInfo):
#     noteId = noteInfo["noteId"]
#     tags: list = noteInfo["tags"]
#     fields = noteInfo["fields"]
#     yuanshu = fields["原书"]["value"]
#     if yuanshu != "" and "english_sentiment" in tags:
#         pqdoc = PyQuery(yuanshu)
#         example_c_entrys = pqdoc("span.example_c")
#         ol_tag = PyQuery("<ol></ol>")
#         ol_tag.add_class("example_english_sentiment")
#         for example_c_entry in example_c_entrys:
#             li_tag = PyQuery("<li></li>")
#             matches = str_en_zh_split(example_c_entry.text)
#             if matches:
#                 li_tag.text(matches[0])
#                 ol_tag.append(li_tag)
#         if len(ol_tag.children()) != 0:
#             # tags.append("english_sentiment")
#             invoke(
#                 "updateNote",
#                 note={
#                     "id": noteId,
#                     "fields": {"英文意境": str(ol_tag)},
#                     # "tags": tags,
#                 },
#             )
#             print(
#                 "es",
#                 f"{index+1}/{len(notesInfo)}",
#                 # ol_tag,
#             )

# >>>
# note_id_list = invoke("findNotes", query="deck:2024红宝书考研词汇")
# notesInfo = invoke("notesInfo", notes=note_id_list)

# mdx = MDX("D:\\Software\\MDictPC\\doc\\Collins\ccald9.mdx")
# headwords = [*mdx]  # 单词名列表
# items = [*mdx.items()]  # 释义html源码列表
# if len(headwords) == len(items):
#     print(f"加载成功：共{len(headwords)}条")
# else:
#     print(f"【ERROR】加载失败{len(headwords)}，{len(items)}")

# for index, noteInfo in enumerate(notesInfo):
#     noteId = noteInfo["noteId"]
#     tags: list = noteInfo["tags"]
#     fields = noteInfo["fields"]
#     words = fields["word"]["value"]
#     if words.encode() in headwords:
#         word, html = items[headwords.index(words.encode())]
#         word, html = word.decode(), html.decode()
#         if "柯林斯" not in tags and str(html) != "":
#             tags.append("柯林斯9")
#             invoke(
#                 "updateNote",
#                 note={
#                     "id": noteId,
#                     "fields": {"柯林斯": str(html)},
#                     "tags": tags,
#                 },
#             )
#             print(fields["word"]["value"], "explain", f"{index+1}/{len(notesInfo)}")
#     else:
#         print(f"【未找到单词】：{words}")

# >>>
# note_id_list = invoke("findNotes", query="deck:2024红宝书考研词汇")
# notesInfo = invoke("notesInfo", notes=note_id_list)

# mdx = MDX("D:/Software/MDictPC/doc/mwa/mwa.mdx")
# headwords = [*mdx]  # 单词名列表
# items = [*mdx.items()]  # 释义html源码列表
# if len(headwords) == len(items):
#     print(f"[TRUE]加载成功：共{len(headwords)}条")
# else:
#     print(f"[ERROR]加载失败{len(headwords)}, {len(items)}")

# # if "pen".encode() in headwords:
# #     word, html = items[headwords.index("pen".encode())]
# #     word, html = word.decode(), html.decode()
# #     pqdoc = PyQuery(html)
# #     explain_entrys = pqdoc(
# #         "div.entry.entry_v2.boxy > div.sblocks > div > div > div > div.sense > div > ul > li:nth-child(-n+2) > div"
# #     )
# #     a = str(explain_entrys)
# #     pass

# for index, noteInfo in enumerate(notesInfo):
#     noteId = noteInfo["noteId"]
#     tags: list = noteInfo["tags"]
#     fields = noteInfo["fields"]
#     words: str = fields["word"]["value"]
#     if words.encode() in headwords:
#         # and "english_sentiment" not in tags
#         word, html = items[headwords.index(words.encode())]
#         word, html = word.decode(), html.decode()
#         # html = html.replace(r"\\", "")
#         pqdoc = PyQuery(html)
#         example_entrys = pqdoc(
#             "div.entry.entry_v2.boxy > div.sblocks > div > div > div > div.sense > div > ul > li:nth-child(-n+1) > div"
#         )
#         ol_tag = PyQuery("<ol></ol>")
#         ol_tag.add_class("english_sentiment")
#         for example_entry in example_entrys:
#             li_tag = PyQuery("<li></li>")
#             li_tag.append(example_entry)
#             ol_tag.append(li_tag)
#         # a = str(ol_tag)
#         if len(ol_tag.children()) != 0:
#             # if "english_sentiment" not in tags:
#             #     tags.append("english_sentiment")
#             invoke(
#                 "updateNote",
#                 note={
#                     "id": noteId,
#                     "fields": {"英文意境": str(ol_tag)},
#                     # "tags": tags,
#                 },
#             )
#             print(
#                 f"explain: {len(example_entrys)}",
#                 fields["word"]["value"],
#                 f"{index+1}/{len(notesInfo)}",
#             )
#     else:
#         print(f"【未找到单词】：{words}")
#         # if "english_sentiment" in tags:
#         #     tags.remove("english_sentiment")
#         #     invoke(
#         #         "updateNote",
#         #         note={
#         #             "id": noteId,
#         #             "tags": tags,
#         #         },
#         #     )


# note_id_list = invoke("findNotes", query="deck:2024红宝书考研词汇")
# notesInfo = invoke("notesInfo", notes=note_id_list)

# for index, noteInfo in enumerate(notesInfo):
#     noteId = noteInfo["noteId"]
#     tags: list = noteInfo["tags"]
#     fields = noteInfo["fields"]
#     words: str = fields["word"]["value"]
#     if fields["柯林斯"]["value"] == "" and "Mdict柯林斯" in tags:
#         tags.remove("Mdict柯林斯")
#         invoke(
#             "updateNote",
#             note={
#                 "id": noteId,
#                 "tags": tags,
#             },
#         )
#         print(fields["word"]["value"], "up tags", f"{index+1}/{len(notesInfo)}")
# if fields["英文意境"]["value"] == "" and "english_sentiment" in tags:
#     tags.remove("english_sentiment")
#     invoke(
#         "updateNote",
#         note={
#             "id": noteId,
#             "tags": tags,
#         },
#     )
#     print(fields["word"]["value"], "up tags", f"{index+1}/{len(notesInfo)}")
# if "explain" in tags:
#     tags.remove("explain")
#     invoke(
#         "updateNote",
#         note={
#             "id": noteId,
#             "tags": tags,
#         },
#     )
#     print(fields["word"]["value"], "up tags", f"{index+1}/{len(notesInfo)}")

# mdx = MDX("D:/Software/MDictPC/doc/ode_glance/ode_glance.mdx")
# headwords = [*mdx]  # 单词名列表
# items = [*mdx.items()]  # 释义html源码列表
# if len(headwords) == len(items):
#     print(f"[TRUE]加载成功：共{len(headwords)}条")
# else:
#     print(f"[ERROR]加载失败{len(headwords)}, {len(items)}")


# note_id_list = invoke("findNotes", query="deck:2024红宝书考研词汇")
# notesInfo = invoke("notesInfo", notes=note_id_list)

# for index, noteInfo in enumerate(notesInfo):
#     noteId = noteInfo["noteId"]
#     tags: list = noteInfo["tags"]
#     fields = noteInfo["fields"]
#     word = fields["word"]["value"]
#     if "english_sentiment" not in tags and "english_sentiment_tidme" not in tags:
#         ol_tag = PyQuery("<ol></ol>")
#         ol_tag.add_class("english_sentiment")
#         exams = download_cambridge_words_exam(word)
#         if len(exams) != 0:
#             for en, zh in exams:
#                 div_tag = PyQuery("<div></div>")
#                 div_tag.add_class("vi_content")
#                 span_tag = PyQuery("<span></span>")
#                 span_tag.add_class("mw_zh")
#                 li_tag = PyQuery("<li></li>")
#                 span_tag.text(zh)
#                 div_tag.html(en + str(span_tag))
#                 li_tag.append(div_tag)
#                 ol_tag.append(li_tag)
#         else:
#             continue
#         a = str(ol_tag)
#         if len(ol_tag.children()) != 0:
#             tags.append("english_sentiment_cambridge")
#             invoke(
#                 "updateNote",
#                 note={
#                     "id": noteId,
#                     "fields": {"英文意境": str(ol_tag)},
#                     "tags": tags,
#                 },
#             )
#             print("es", f"{index+1}/{len(notesInfo)}")


# ——————————————————————
# note_id_list = invoke("findNotes", query="deck:2024红宝书考研词汇")
# notesInfo = invoke("notesInfo", notes=note_id_list)

# backup = dict()

# for index, noteInfo in enumerate(notesInfo):
#     noteId = noteInfo["noteId"]
#     tags: list = noteInfo["tags"]
#     fields = noteInfo["fields"]
#     word = fields["word"]["value"]
#     aabbyy = fields["意境嵌入词义"]["value"]
#     if aabbyy and aabbyy != "":
#         aabbyy_doc = PyQuery(aabbyy)
#         c = str(aabbyy_doc)
#         li_list = aabbyy_doc.children()
#         if len(li_list) > 4:
#             li_list = backward_erasure(li_list, 4)
#             ol_tag = PyQuery("<ol></ol>")
#             ol_tag.add_class("english_sentiment")
#             for li in li_list:
#                 ol_tag.append(li)
#             # 更新
#             m = str(ol_tag)
#             if len(ol_tag.children()) != 0:
#                 invoke(
#                     "updateNote",
#                     note={
#                         "id": noteId,
#                         "fields": {"意境嵌入词义": str(ol_tag)},
#                     },
#                 )
#                 print("es", f"{index+1}/{len(notesInfo)}")

# with open("./backup.json", "rb") as f:
#     raw_data = f.read()
#     json_data = json.loads(raw_data)

# for index, noteInfo in enumerate(notesInfo):
#     noteId = noteInfo["noteId"]
#     tags: list = noteInfo["tags"]
#     fields = noteInfo["fields"]
#     word = fields["word"]["value"]
#     aabbyy = fields["意境嵌入词义"]["value"]
#     ab = json_data[word]
#     try:
#         invoke(
#             "updateNote",
#             note={
#                 "id": noteId,
#                 "fields": {"意境嵌入词义": ab},
#             },
#         )
#         print("es", f"{index+1}/{len(notesInfo)}")
#     except Exception as e:
#         print(e)

#     if not aabbyy == "":
#         backup.update({word: aabbyy})
# backup_bytes = json.dumps(backup, ensure_ascii=False).encode()
# with open("./backup.json", "wb") as f:
#     f.write(backup_bytes)


# ——————————————————————

html = """
☀ <span class="category">**国家 国土：**</span>
<span class="vocabulary">**country**</span> ['kʌntrɪ] 
<span class="definition">n. [C] 国家、国。是此义的基础用词，使用范围最广：</span>beautiful, great country 美丽的国家；伟大的国家 / developed country 发达国家 / home, native, own country 祖国 / host country 东道国 / neighbouring/neighboring country 邻国 / sovereign country（尤美）主权国家 / neutral country 中立国 / Western country 西方国家 / English-speaking country 讲英语的国家 / member country 成员国 / socialist country 社会主义国家 / I love my country. 我爱我的祖国。/ The two countries signed a basic treaty of cooperation. 两国签署了一项基本的合作条约。/ Over 30 countries participated in the Games. 有30多个国家参加了这次运动会。
"""

# 使用BeautifulSoup解析HTML
soup = BeautifulSoup(html, "lxml")

# 找到特定的元素
link1 = soup.find("span", class_="category")

# 获取该元素之后的裸文本
# .next_sibling可以获得元素的下一个兄弟节点
# .strip()用于去除字符串首尾的空白字符
# 如果下一个兄弟节点不是文本节点，则递归查找直到找到文本节点为止
# bare_text_after_link1 = link1.next_sibling.strip()
# while bare_text_after_link1 and not bare_text_after_link1.strip():
#     bare_text_after_link1 = bare_text_after_link1.next_sibling.strip()

# print(bare_text_after_link1)

# ——————————————————————


# note_id_list = invoke("findNotes", query="deck:24意境语义红宝石")
# notesInfo = invoke("notesInfo", notes=note_id_list)

# headwords, items = u_mdx("D:/Software/MDictPC/doc/ode_glance/ode_glance.mdx")
# for index, noteInfo in enumerate(notesInfo):
#     noteId = noteInfo["noteId"]
#     tags: list = noteInfo["tags"]
#     fields = noteInfo["fields"]
#     anki_word = fields["word"]["value"]
# if anki_word.encode() in headwords:
#     word, html = items[headwords.index(anki_word.encode())]
#     word, html = word.decode(), html.decode()
#     pq = PyQuery(html)
#     exampleGroup = pq(".exampleGroup.exGrBreak")
#     # 查找有翻译的句子。
#     first_example = None
#     first_example_sentence_en = ""
#     first_example_sentence_cn = ""
#     # 有中文的
#     for item in exampleGroup:
#         child = PyQuery(item).find('.cn').eq(0)
#         if child:
#             first_example = PyQuery(item)
#             first_example_sentence_en = first_example(".example").html()
#             first_example_sentence_cn = first_example(".cn").html()
#             first_example_sentence_en = first_example_sentence_en.replace(anki_word,
#                                                                           f"<span class='decorated_word'>{anki_word}</span>")
#             break
#     # 否则，随便找个有英文的
#     if not first_example:
#         for item in exampleGroup:
#             child = PyQuery(item).find('.example').eq(0)
#             if child:
#                 first_example = PyQuery(item)
#                 first_example_sentence_en = first_example(".example").html()
#                 first_example_sentence_en = first_example_sentence_en.replace(anki_word,
#                                                                               f"<span class='decorated_word'>{anki_word}</span>")
#                 break
#
#     invoke(
#         "updateNote",
#         note={
#             "id": noteId,
#             "fields": {"q_sentence": first_example_sentence_en, "a_sentence": first_example_sentence_cn},
#             "tag": []
#         },
#     )
#     print(f"{index + 1}/{len(notesInfo)}")


# for index, noteInfo in enumerate(notesInfo):
#     noteId = noteInfo["noteId"]
#     tags: list = noteInfo["tags"]
#     fields = noteInfo["fields"]
#     try:
#         SA: dict = json.loads(fields["SA"]["value"].replace("'", '"'))
#         for key in SA.keys():
#             SA[key] = SA[key][:6]
#         result = json.dumps(SA, ensure_ascii=False)
#         invoke(
#             "updateNote",
#             note={"id": noteId, "fields": {"SA": result}, "tags": []},
#         )
#         print(fields["word"]["value"], f"{index + 1}/{len(notesInfo)}")

#     except Exception as e:
#         print(e)
# ——————————————————————


# nlp = spacy.load("en_core_web_trf")
# # nlp = spacy.load("en_core_web_sm")
# # 创建 Porter 词干提取器
# stemmer = PorterStemmer()


# def jaccard_similarity(set1, set2):
#     intersection = len(set1 & set2)
#     union = len(set1 | set2)
#     return intersection / union


# word1 = "apple"
# word2 = "appel"


# def cmp_2word(word1, word2):
#     if word1 == word2:
#         return True
#     if jaccard_similarity(set(word1), set(word2)) < 0.76:
#         return False
#     else:
#         # 剩下百分比[80, 100)交给它处理。
#         doc1 = nlp(word1)
#         doc2 = nlp(word2)
#         # 先进行词形还原，若不行再进行词干提取
#         if doc1[0].lemma_ == doc2[0].lemma_:
#             return True
#         elif stemmer.stem(doc1[0].text) == stemmer.stem(doc2[0].text):
#             return True
#         else:
#             return False


# res = cmp_2word("impress", "impressed")
# pass


# def request(action, **params):
#     return {"action": action, "params": params, "version": 6}


# def invoke(action, **params):
#     requestJson = json.dumps(request(action, **params)).encode("utf-8")
#     response = json.load(
#         urllib.request.urlopen(
#             urllib.request.Request("http://127.0.0.1:8765", requestJson)
#         )
#     )
#     if len(response) != 2:
#         raise Exception("response has an unexpected number of fields")
#     if "error" not in response:
#         raise Exception("response is missing required error field")
#     if "result" not in response:
#         raise Exception("response is missing required result field")
#     if response["error"] is not None:
#         raise Exception(response["error"])
#     return response["result"]


# with open("./backup.json", "rb") as f:
#     raw_data = f.read()
#     json_data = json.loads(raw_data)


# note_id_list = invoke("findNotes", query="deck:2024红宝书考研词汇")
# notesInfo = invoke("notesInfo", notes=note_id_list)

# for index, noteInfo in enumerate(notesInfo):
#     noteId = noteInfo["noteId"]
#     tags: list = noteInfo["tags"]
#     fields = noteInfo["fields"]
#     word = fields["word"]["value"]
#     aabbyy = fields["意境嵌入词义"]["value"]
#     if not aabbyy or aabbyy == "":
#         for item in json_data.values():
#             # 处理文本
#             aabbyy_doc = PyQuery(item).children()
#             for li_tag in aabbyy_doc:
#                 mli = PyQuery(li_tag)
#                 mli("span.mw_zh").remove()
#                 a = str(mli)
#                 doc = nlp(html.unescape(mli.text()))
#                 filtered_sentence = [
#                     token.text
#                     for token in doc
#                     if not nlp.vocab[token.text].is_stop
#                     and not nlp.vocab[token.text].is_punct
#                 ]

#                 # 遍历文档中的词并打印词形还原形式
#                 for token in filtered_sentence:
#                     if cmp_2word(word, token):
#                         # 说明句子中有目标单词。
#                         # 不可行时间太长。
#                         b = str(li_tag)
#                         pass

#         print(f"{index+1}/{len(notesInfo)}")

# wb = openpyxl.load_workbook("C:/Users/Snowy/Desktop/工作簿1.xlsx")
# ws = wb.active
# cell_range = [
#     (f"sentence {cell1.value}", cell2.value, cell3.value, cell4.value.split("、"))
#     for (cell1, cell2, cell3, cell4) in ws["A2":"D21"]
# ]


# # nlp = spacy.load("en_core_web_trf")
# nlp = spacy.load("en_core_web_sm")
# paragraph = ws["B2"].value
# doc = nlp(paragraph)
# # 获取分词结果

# filtered_sentence = [
#     token.text
#     for token in doc
#     if not nlp.vocab[token.text].is_stop and not nlp.vocab[token.text].is_punct
# ]

# print(filtered_sentence)


# pass
