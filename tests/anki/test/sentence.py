import openpyxl
import spacy  # 导包
from nltk.stem import PorterStemmer

wb = openpyxl.load_workbook("C:/Users/Snowy/Desktop/工作簿1.xlsx")
ws = wb.active
cell_range = [
    (f"sentence {cell1.value}", cell2.value, cell3.value, cell4.value.split("、"))
    for (cell1, cell2, cell3, cell4) in ws["A2":"D21"]
]


# nlp = spacy.load("en_core_web_trf")
nlp = spacy.load("en_core_web_sm")
paragraph = ws["B2"].value
doc = nlp(paragraph)
# 获取分词结果

filtered_sentence = [
    token.text
    for token in doc
    if not nlp.vocab[token.text].is_stop and not nlp.vocab[token.text].is_punct
]

print(filtered_sentence)


pass
